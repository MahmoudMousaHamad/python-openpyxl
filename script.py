from openpyxl import load_workbook
import re
import base64
import uuid

class Pricing:
    def __init__(self, code, encoded_code, title, kind, price):
        self.code = code
        self.encoded_code = encoded_code
        self.title = title
        self.kind = kind
        self.price = price
    
    def generate_anchor_element(self):
        if self.kind == "CEO":
            return f'<a href="/entry-form-executives/?order_id={uuid.uuid4()}&sub_sector={self.title}&code={self.encoded_code}">{self.title}</a>'
        else:
            return f'<a href="/entry-form-organizations/?order_id={uuid.uuid4()}&sub_sector={self.title}&code={self.encoded_code}">{self.title}</a>'

    def generate_case_statement(self):
        return f"case '{self.code}': price = {self.price}; break;"



wb = load_workbook("Workbook.xlsx")

sheet = wb.active

max_row = sheet.max_row
max_col = sheet.max_column

pricings = {}

current_sector = ""

# row
for i in range(1, max_row + 1):
    row_data = []
    is_pricing_row = False
    

    # column
    for j in range(1, max_col + 1):
        cell = sheet.cell(row = i, column = j)
        cell_val = cell.value
        if cell_val:
            if "Sector" in cell_val and "CEO" not in cell_val and "Chairman" not in cell_val:
                current_sector = cell_val
                pass
            regex = re.compile(r'^[A-Z][0-9]{1,2}?$')
            if regex.search(cell_val):
                is_pricing_row = True
            if is_pricing_row:
                row_data.append(cell_val.strip())
            print(cell_val.strip(), end = " ")
    if is_pricing_row:
        code = row_data[0]
        encoded_code = base64.b64encode(code.encode("utf-8"))
        title = row_data[1]
        kind = 'CEO' if 'CEO' in title else 'ORG'
        price = row_data[2][3:]
        pricing = Pricing(code, encoded_code, title, kind, price)
        if current_sector in pricings:
            pricings[current_sector].append(pricing)
        else:
            pricings[current_sector] = []
            pricings[current_sector].append(pricing)

        print(pricing.generate_anchor_element())
    print('\n')

# Write to files
fa = open("anchor_elements.html", "w")
fa.close()

fc = open("case_statements.txt", "w")
fc.close()

fa = open("anchor_elements.html", "a")
fc = open("case_statements.txt", "a")

fa.write("<div class='et_pb_module et_pb_accordion et_pb_accordion_0'>\n")

elemets_counter = 0

for sector, pricings_list in pricings.items():
    elemets_counter += 1
    fa.write(f"<div class='et_pb_toggle et_pb_module et_pb_accordion_item et_pb_accordion_item_{elemets_counter}  et_pb_toggle_close'>\n")
    fa.write(f"<h5 class='et_pb_toggle_title'>{sector}</h5>\n")
    fa.write("<div class='et_pb_toggle_content clearfix'>\n")
    for i in range(0, len(pricings_list)):
        fa.write("<div>\n")
        fa.write(pricings_list[i].generate_anchor_element() + '\n')
        fa.write("</div>\n")
        fc.write(pricings_list[i].generate_case_statement())
    fa.write("</div>\n")
    fa.write("</div>\n")
    #fa.write(pricings[i].generate_anchor_element() + '\n')
    #fc.write(pricings[i].generate_case_statement() + '\n')

fa.write("</div>\n")
fa.close()
fc.close()
